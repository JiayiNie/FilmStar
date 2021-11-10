from openpyxl import load_workbook
from openpyxl.styles import Font
a = True
while a:
    print("-------------------------------------------------------------------------------")
    print("This program will generate F_Design, B_Design, T_Design")
    print("-------------------------------------------------------------------------------")
    fname = str(input("Please input the exact name of the file you want to process: "))
    try:
        f = open(fname)
        a = False
        
    except FileNotFoundError:
        print('File does not exist')
print("hello")
print(fname)

workbook = load_workbook(filename = fname)
print("bye")
print(workbook.sheetnames)
sheet = workbook["F_Design"]
workbook.remove(sheet)
sheet = workbook["B_Design"]
workbook.remove(sheet)
sheet = workbook["T_Design"]
workbook.remove(sheet)
print(workbook.sheetnames)
workbook.save(filename = fname)