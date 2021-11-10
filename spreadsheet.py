from openpyxl import load_workbook
from openpyxl.styles import Font


def get_CR(sheet):
    mylist = []
    for x in range(len(notes_Col)):
        if "{CR}" in notes_Col[x].value:
            mylist.append(x+1)
    return mylist

def change_A(places,sheet,design_sheet):
    ft = Font(color = "FF0000")
    A_Col = sheet["I"]
    for i in range(len(places)):
        j = places[i]
        A_value = (A_Col[j-1].value+A_Col[j-2].value)/2
        cell = design_sheet["I%d" % (j)]
        cell.font = ft
        cell.value = A_value

def get_NR(sheet):
    mylist = []
    for x in range(len(notes_Col)):
        if "{NR}" in notes_Col[x].value:
            mylist.append(x+1)
    return mylist

def change_B(places,sheet,design_sheet):
    ft = Font(color = "FF0000")
    B_Col = sheet["J"]
    for i in range(len(places)):
        j = places[i]
        B_value = (B_Col[j-1].value+B_Col[j-2].value)/2
        cell = design_sheet["J%d" % (j)]
        cell.font = ft
        cell.value = B_value

def get_PR(sheet):
    mylist = []
    for x in range(len(notes_Col)):
        if "{PR}" in notes_Col[x].value:
            mylist.append(x+1)
    return mylist

def change_C(places,sheet,design_sheet):
    ft = Font(color = "FF0000")
    C_Col = sheet["K"]
    for i in range(len(places)):
        j = places[i]
        if C_Col[j-1].value != None and C_Col[j-2].value != None:
            C_value = (C_Col[j-1].value+C_Col[j-2].value)/2
            cell = design_sheet["K%d" % (j)]
            cell.font = ft
            cell.value = C_value

def delete_QW_B(sheet):
    mylist = []
    for x in range(len(notes_Col)):
        if "{QW}" in notes_Col[x].value:
            mylist.append(x+1)
    sheet.delete_rows(mylist[0],len(notes_Col)-mylist[0]+1)
    cell_A = sheet["I2"]
    cell_Incident = sheet["Q2"]
    cell_A.value = 3.552
    cell_Incident.value = "Incident from MQW"

def delete_QW_T(sheet,fname):
    mylist = []
    for x in range(len(notes_Col)):
        if "{QW}" in notes_Col[x].value:
            mylist.append(x+1)
    # print(mylist)
    last_row = mylist[-1]
    sheet.delete_rows(4, last_row-4+1)

    workbook.save(filename = fname)
    cell_A = sheet["I2"]
    cell_Incident = sheet["Q2"]
    cell_A.value = 3.552
    cell_Incident.value = "Incident from MQW"

    cell_A1 = sheet["I3"]
    cell_Incident1 = sheet["Q3"]
    cell_A1.value = 1
    cell_Incident1.value = "Air substrate"

    cell_A2 = sheet["J3"]
    cell_A2.value = 0

    cell_A2 = sheet["H3"]
    cell_A2.value = None

def change_repeat(sheet):
    num_col = sheet["C"]
    repeat_col = sheet["D"]

    num_list = []
    repeat_list = []

    for j in range(3,len(repeat_col)):
        repeat_list.append(repeat_col[j].value)
    for k in range(3,len(num_col)):
        num_list.append(num_col[k].value)
    
    index = 0
    while index < len(num_list):
        if num_list[index] != None:
            
            cell_num = sheet["C%d"%(index-int(num_list[index])+1+4)]
            # print("  "+ str(index-int(num_list[index])+1+3))
            cell_repeat = sheet["D%d"%(index-int(num_list[index])+1+4)]
            cell_n = sheet["C%d"%(int(index)+4)]
            cell_r = sheet["D%d"%(int(index)+4)]

            cell_num.value = num_list[index]
            cell_repeat.value = repeat_list[index]
            cell_n.value = None
            cell_r.value = None
            
        index += 1


if __name__ == "__main__":
    a = True
    # while a:
    #     print("-------------------------------------------------------------------------------")
    #     print("This program will generate F_Design, B_Design, T_Design")
    #     print("-------------------------------------------------------------------------------")
    #     fname = str(input("Please input the exact name of the file you want to process: "))
    #     try:
    #         f = open(fname)
    #         a = False
            
    #     except FileNotFoundError:
    #         print('File does not exist')
    # print("hello")
    # print(fname)
    fname = "analyze_S940 E01 S.xlsx"
    workbook = load_workbook(filename = fname)
    print("bye")
    print(workbook.sheetnames)
    sheet = workbook["F_Design"]
    workbook.remove(sheet)
    sheet = workbook["B_Design"]
    workbook.remove(sheet)
    sheet = workbook["T_Design"]
    workbook.remove(sheet)
    print(workbook.sheetnames)
    workbook.save(filename = fname)
    raw_sheet = workbook.active
    raw_sheet = workbook["Design Inputs"]

    workbook.copy_worksheet(raw_sheet)

    F_Design = workbook["Design Inputs Copy"]
    F_Design.title = "F_Design"
    
    notes_Col = raw_sheet["Q"]
    
    a = get_CR(raw_sheet)
    # print(a)
    change_A(a, raw_sheet, F_Design)

    b = get_NR(raw_sheet)
    # print(b)
    change_B(b, raw_sheet, F_Design)

    c = get_PR(raw_sheet)
    # print(c)
    change_C(c, raw_sheet, F_Design)

    workbook.save(filename = fname)

    workbook.copy_worksheet(F_Design)
    B_Design = workbook["F_Design Copy"]
    B_Design.title = "B_Design"

    delete_QW_B(B_Design)

    workbook.save(filename = fname)


    workbook.copy_worksheet(F_Design)
    T_Design = workbook["F_Design Copy"]
    T_Design.title = "T_Design"

    delete_QW_T(T_Design,fname)
    workbook.save(filename = fname)

    
    my_row_list = []
    row = T_Design.max_row
    # print("maxrow"+str(row))
    i = 4
    for row in range(4, row+1):
        r_list = []
        r = T_Design["%d" %i]
        for j in r:
            r_list.append(j.value)
        my_row_list.append(r_list)
        i+=1
    
    for y in range(0, len(my_row_list)):
        x = 0
        w = T_Design["%d"%row]
        for z in w:
            z.value = my_row_list[y][x]
            x+=1
        row-=1
    
    workbook.save(filename = fname)

    change_repeat(T_Design)
    workbook.save(filename = fname)
    # print(workbook.sheetnames)
    print("WORK DONE")
    print(workbook.sheetnames)
    f.close()